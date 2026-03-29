from openpyxl import load_workbook

COLUMN_ALIASES = {
    "question": "prompt_text",
    "prompt": "prompt_text",
    "문제": "prompt_text",
    "질문": "prompt_text",
    "answer": "template_answer",
    "template": "template_answer",
    "답변": "template_answer",
    "템플릿": "template_answer",
    "정답": "template_answer",
    "type": "type",
    "유형": "type",
    "타입": "type",
    "topic": "topic",
    "주제": "topic",
    "sequence": "sequence",
    "순서": "sequence",
    "bullet_points": "bullet_points",
    "bullets": "bullet_points",
    "professor_prompt": "professor_prompt",
    "교수": "professor_prompt",
    "student_response_1": "student_response_1",
    "학생1": "student_response_1",
    "student_response_2": "student_response_2",
    "학생2": "student_response_2",
}


def parse_xlsx(file_path: str) -> list[dict]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row and map to standard column names
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = []
    for h in raw_headers:
        if h is None:
            headers.append(None)
            continue
        key = str(h).strip().lower().replace(" ", "_")
        headers.append(COLUMN_ALIASES.get(key, key))

    if "prompt_text" not in headers:
        wb.close()
        raise ValueError("필수 컬럼 'prompt_text'을 찾을 수 없습니다. 컬럼명을 확인해주세요.")

    # Read data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = val if val is not None else None
        if not record.get("prompt_text"):
            continue
        if not record.get("type"):
            record["type"] = "speaking_interview"
        record["prompt_text"] = str(record["prompt_text"]).strip()
        if record.get("template_answer"):
            record["template_answer"] = str(record["template_answer"]).strip()
        rows.append(record)

    wb.close()
    return rows
